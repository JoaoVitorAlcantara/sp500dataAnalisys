#Imports

from csv import writer
from turtle import color
from openpyxl import Workbook
import pandas as pd
from yahooquery import Ticker 
from time import sleep
from datetime import date, timedelta
import openpyxl as excel
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image


#Preparing

stocks = pd.read_excel('S&P_500.xlsx')
stocks = stocks['Symbol']
 
tickers = dict()
tickers_order = list()
tickers_errors = list()
today = date.today()
yesterday = today - timedelta(days=1)

#Getting the information online


#Stocks

for t in stocks:
    try:
        info = Ticker(t).price
        percent = round(float((info[f'{t}']['regularMarketPrice']-info[f'{t}']['regularMarketOpen'])/info[f'{t}']['regularMarketOpen']), 4)
        var_ticker = {'Code': t, 'Apreciation': percent}
        tickers[f'{t}']= var_ticker
        tickers_order.append(percent)
        print(f'Ticker {t} calculated')
        print(tickers_order)
    except:
        tickers_errors.append(t)
        print(f'Ticker {t} not found')
    print(f'{len(stocks)-len(tickers)} tickers left')

#S&P500

info = Ticker('^GSPC').price
info_sp = round((info['^GSPC']['regularMarketPrice']-info['^GSPC']['regularMarketOpen'])/info['^GSPC']['regularMarketOpen'], 4)

#Ordering Top

apreciation_order_top = tickers_order
apreciation_order_top.sort(reverse=True)
apreciation_order_top = apreciation_order_top[0:5]

#Ordering Low

apreciation_order_low = tickers_order
apreciation_order_low.sort()
apreciation_order_low = apreciation_order_low[0:5]


code_order_top = list()
code_order_low = list()

#Matching the variations to the tickers

for tic in tickers:
    for aprec in apreciation_order_top:
        if tickers[f'{tic}']['Apreciation'] == aprec:
            code_order_top.append(tic)
    for aprec in apreciation_order_low:
        if tickers[f'{tic}']['Apreciation'] == aprec:
            code_order_low.append(tic)            

#Adding S&P500's Info

code_order_top.append('S&P500')
code_order_low.append('S&P500')
apreciation_order_top.append(info_sp)
apreciation_order_low.append(info_sp)

#Building graphic 1

table = pd.DataFrame(list(zip(code_order_top, apreciation_order_top)),
              columns=['Top 5','Variation'])

table.plot(x = 'Top 5', y = 'Variation', kind = 'bar', color='black')
plt.tight_layout()
plt.savefig('grafico_top.png')
plt.close()

#Building graphic 2

table2 = pd.DataFrame(list(zip(code_order_low, apreciation_order_low)),
              columns=['Low 5','Variation'])

table2.plot(x = 'Low 5', y = 'Variation', kind = 'bar', color='black')
plt.tight_layout()
plt.savefig('grafico_low.png')
plt.close()

#Creating Excel file

#Adding Dataframes

writer = pd.ExcelWriter('Results.xlsx', engine='openpyxl')
table.to_excel(writer, sheet_name='Top', index=False)
table2.to_excel(writer, sheet_name='Low', index=False)

writer.save()

#Adding images

archive = excel.load_workbook('Results.xlsx')
archive.create_sheet('Graphic_1')
archive.create_sheet('Graphic_2')

page2 = archive['Graphic_1']
img1 = Image('grafico_top.png')
page2.add_image(img1, "F2")

page4 = archive['Graphic_2']
img2 = Image('grafico_low.png')
page4.add_image(img2, "F2")



archive.save("Results.xlsx")